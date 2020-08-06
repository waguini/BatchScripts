import openpyxl
import sys
from pathlib import Path


#pega os argumentos 
demanda = sys.argv[1]
jira = sys.argv[2]
sistema = sys.argv[3]
autor = sys.argv[4]
urlgit = sys.argv[5]


rec = jira.split('-')
recnum = rec[0]
jiranum = rec[1]

# Monta os diretorios
diretorio_atual = Path().parent.absolute()
PATH_GERAL = diretorio_atual / demanda / jira
# converte pra string pra poder passar os parametros no nome dos diretorios
pathStr = str(PATH_GERAL) 

PATH_ENTRY_CRITERIA = str(pathStr + '\\01 - Entry Criteria\\TLF - Entry Criteria - ['+jira+']_TEMPLATE.xlsx')
PATH_EXIT_CRITERIA = str(pathStr + '\\04 - Exit Criteria\\TLF - Exit Criteria - ['+jira+']_TEMPLATE.xlsx')
PATH_MATRIZ_RATREABILIDADE = str(pathStr + '\\06 - Matriz de Rastreabilidade\\TLF - Matriz de Rastreabilidade - ['+jira+']_TEMPLATE.xlsx')



# ================ ENTRY CRITERIA =========================== #
print ('Alterando o Entry Criteria...')
wb = openpyxl.load_workbook(PATH_ENTRY_CRITERIA)
sheet = wb['Critérios']

entrycriteria = 'ENTRY CRITERIA - PJ '+ recnum + ' - '+ demanda
sheet['B2'].value = entrycriteria

# coloca o Jira no user story
sheet['F5'].value = jiranum

wb.save(PATH_ENTRY_CRITERIA)
wb.close
# ================ FIM ENTRY CRITERIA ======================= #


# ================ EXIT CRITERIA ============================ #
print ('Alterando o Exit Criteria...')
wb = openpyxl.load_workbook(PATH_EXIT_CRITERIA)
exitcriteria = '   EXIT CRITERIA '+ recnum+ ': '+ demanda
sheet = wb['Exit Criteria']	
sheet['A2'].value = exitcriteria

# limpa o campo da documentação
sheet['B5'].value = ''

# preenche o campo sistemas trabalhados
sheet['B4'].value = sistema

# Preenche o campo Autor
sheet['B6'].value = autor

wb.save(PATH_EXIT_CRITERIA)
wb.close
# ================ FIM EXIT CRITERIA ============================ #

# ================ MATRIZ DE RASTREABILIDADE ==================== #
print ('Alterando a Matriz de Rastreabilidade...')
wb = openpyxl.load_workbook(PATH_MATRIZ_RATREABILIDADE)
matriz = '   Matriz de Rastreabilidade '+recnum+': '+ demanda

sheet = wb['Matriz de Rastreabilidade']
sheet['A2'].value = matriz
sheet['B4'].value = sistema
sheet['B5'].value = urlgit

wb.save(PATH_MATRIZ_RATREABILIDADE)
wb.close

# ================ FIM MATRIZ DE RASTREABILIDADE ================ #